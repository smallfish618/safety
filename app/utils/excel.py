import openpyxl
from app import db
from app.models.station import FireStation, EquipmentExpiry, ResponsiblePerson
from datetime import datetime

def import_excel_data(file_paths):
    """从Excel文件导入数据到数据库"""
    
    # 导入微型消防站物资表
    wb = openpyxl.load_workbook(file_paths['station'])
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2):  # 假设第一行是表头
        if not row[0].value:  # 如果设备区域编码为空，跳过
            continue
            
        # 处理日期格式，如果是字符串需要转换成日期对象
        production_date = row[6].value
        if isinstance(production_date, str):
            try:
                production_date = datetime.strptime(production_date, '%Y-%m-%d').date()
            except ValueError:
                production_date = None
        
        station = FireStation(
            area_code=row[0].value,
            area_name=row[1].value,
            item_name=row[2].value,
            manufacturer=row[3].value,
            model=row[4].value,
            quantity=row[5].value,
            production_date=production_date,
            certificate=row[7].value,
            certificate_no=row[8].value,
            remark=row[9].value if len(row) > 9 else None
        )
        db.session.add(station)
    
    # 导入设备有效期表
    wb = openpyxl.load_workbook(file_paths['expiry'])
    sheet = wb.active
        
    for row in sheet.iter_rows(min_row=2):  # 假设第一行是表头
        if not row[0].value:  # 如果物资分类为空，跳过
            continue
                
        expiry = EquipmentExpiry(
            item_category=row[0].value,
            item_name=row[1].value,
            normal_expiry=float(row[2].value) if row[2].value is not None else None,
            mandatory_expiry=float(row[3].value) if row[3].value is not None else None,
            description=row[4].value if len(row) > 4 else None
        )
        db.session.add(expiry)

   
    # 导入负责人表
    wb = openpyxl.load_workbook(file_paths['responsible'])
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2):  # 假设第一行是表头
        if not row[0].value:  # 如果设备区域编码为空，跳过
            continue
            
        person = ResponsiblePerson(
            area_code=row[0].value,
            area_name=row[1].value,
            person_name=row[2].value,
            contact=row[3].value,
            email=row[4].value if len(row) > 4 else None
        )
        db.session.add(person)
    
    db.session.commit()
    return True

