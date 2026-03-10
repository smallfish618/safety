"""
生成数据导入Excel模板
包含消防器材表和微型消防站表两个Sheet
自动包含有效的区域编码列表
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 导入获取区域编码的函数
sys.path.insert(0, '.')
from get_valid_area_codes import get_area_codes

def create_template():
    """创建Excel模板"""
    
    # 获取有效的区域编码
    try:
        area_codes = get_area_codes()
    except Exception as e:
        print(f"[WARN] Failed to fetch area codes: {str(e)}")
        area_codes = []
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的Sheet
    wb.remove(wb.active)
    
    # 创建第一个Sheet - 消防器材
    ws_equipment = wb.create_sheet('消防器材')
    equipment_headers = [
        '区域编码',
        '区域名称',
        '楼层',
        '安装位置',
        '器材类型',
        '器材名称',
        '型号',
        '重量',
        '数量',
        '生产日期',
        '使用年限',
        '到期日期',
        '备注'
    ]
    
    ws_equipment.append(equipment_headers)
    
    # 添加示例数据
    sample_equipment = [
        ['001', '电气室', '1楼', '走廊', '火灾报警器', 'ABX型烟感探测器', 'ABX-SD-001', '0.1kg', 5, '2023-01-15', '8年', '2031-01-15', '需要定期检测'],
        ['002', '配电室', '地下1楼', '配电柜上方', '灭火器', 'ABC干粉灭火器', 'MFZ/8-1', '8kg', 3, '2022-06-20', '5年', '2027-06-20', ''],
    ]
    
    for row in sample_equipment:
        ws_equipment.append(row)
    
    # 创建第二个Sheet - 微型消防站
    ws_station = wb.create_sheet('微型消防站')
    station_headers = [
        '区域编码',
        '区域名称',
        '物品名称',
        '生产厂家',
        '型号',
        '数量',
        '生产日期',
        '合格证',
        '合格证编号',
        '备注'
    ]
    
    ws_station.append(station_headers)
    
    # 添加示例数据
    sample_station = [
        ['001', '电气室', '灭火器', '青岛消防器材有限公司', 'ABC干粉8kg', '2', '2023-05-10', '有', 'CCC123456', '置于明显位置'],
        ['002', '配电室', '应急照明灯', '飞利浦', 'EL500W', '4', '2023-01-20', '有', 'CCC654321', ''],
    ]
    
    for row in sample_station:
        ws_station.append(row)
    
    # 设置样式
    for ws in [ws_equipment, ws_station]:
        # 字体和填充
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 设置数据行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 设置列宽
        for col_num, header in enumerate(ws[1], 1):
            column_letter = get_column_letter(col_num)
            if col_num in [2, 3, 4, 6, 7, 8, 9, 10]:  # 调整特定列的宽度
                ws.column_dimensions[column_letter].width = 18
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
    
    # 创建说明Sheet
    ws_readme = wb.create_sheet('使用说明', 0)
    
    readme_content = [
        ['消防器材数据导入模板使用说明'],
        [],
        ['表情况1：消防器材表'],
        ['字段说明：'],
        ['• 区域编码：数字，唯一标识，必填'],
        ['• 区域名称：文本，建筑区域名称，必填'],
        ['• 楼层：文本，如"1楼"、"地下1楼"，必填'],
        ['• 安装位置：文本，具体安装地点，必填'],
        ['• 器材类型：文本，如"灭火器"、"火灾报警器"，必填'],
        ['• 器材名称：文本，具体器材名称，必填'],
        ['• 型号：文本，品牌和型号，必填'],
        ['• 重量：文本，如"8kg"，必填'],
        ['• 数量：数字，整数，可选'],
        ['• 生产日期：日期格式(YYYY-MM-DD)，可选'],
        ['• 使用年限：文本，如"5年"、"8年"，必填'],
        ['• 到期日期：日期格式(YYYY-MM-DD)，必填'],
        ['• 备注：文本，其他说明信息，可选'],
        [],
        ['表情况2：微型消防站表'],
        ['字段说明：'],
        ['• 区域编码：数字，唯一标识，必填'],
        ['• 区域名称：文本，建筑区域名称，必填'],
        ['• 物品名称：文本，物品名称，必填'],
        ['• 生产厂家：文本，生产厂家名称，可选'],
        ['• 型号：文本，规格型号，可选'],
        ['• 数量：文本，如"2件"、"4组"，可选'],
        ['• 生产日期：日期格式(YYYY-MM-DD)，可选'],
        ['• 合格证：文本，"有"或"无"，可选'],
        ['• 合格证编号：文本，证号，可选'],
        ['• 备注：文本，其他说明，可选'],
        [],
        ['有效的区域编码列表：'],
        ['注意：区域编码和名称必须与下列列表匹配！'],
        [],
    ]
    
    # 添加有效的区域编码表（如果有）
    if area_codes:
        readme_content.append(['区域编码', '区域名称'])
        for code, name in area_codes:
            readme_content.append([code, name])
    else:
        readme_content.append(['[无可用区域编码 - 请先在系统中设置负责人信息]'])
    
    readme_content.extend([
        [],
        ['注意事项：'],
        ['1. 日期格式统一为 YYYY-MM-DD (如 2023-01-15)'],
        ['2. 区域编码【必须】来自上面的"有效的区域编码列表"'],
        ['3. 不要删除Sheet页签，只需修改数据'],
        ['4. 不要修改列的顺序和列名'],
        ['5. 可删除示例数据行，从第2行开始填写新数据'],
        ['6. 必填项不能为空，否则导入会失败'],
        ['7. 导入前请先执行备份操作'],
        ['8. 导入后系统会验证区域编码是否有效'],
    ])
    
    for row in readme_content:
        ws_readme.append(row)
    
    # 设置说明Sheet的样式
    for col in ws_readme.columns:
        if col:
            col[0].column_letter
    ws_readme.column_dimensions['A'].width = 60
    
    cell = ws_readme['A1']
    cell.font = Font(bold=True, size=14)
    
    # 保存文件
    output_path = 'e:/safety/data/batch_update_template.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"[OK] Excel模板已创建: {output_path}")
    print(f"[OK] Sheet页签:")
    print(f"  1. 使用说明 - 查看字段说明和注意事项")
    print(f"  2. 消防器材 - 填写消防器材数据")
    print(f"  3. 微型消防站 - 填写微型消防站数据")


if __name__ == '__main__':
    create_template()
